"""
XLReader: A utility to convert Excel spreadsheets containing pallet records
to a more usable format.
Using PyQt5 for its GUI and openpyxl for file parsing.

Written by Benjamin Steenhoek 6/30/2017.

Parses an Excel workbook and compiles all unique records, then saves the result to a new workbook.
"""
#!\C:\Python\python

import sys
import openpyxl

from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLineEdit, QLabel, QFileDialog, QProgressBar, QVBoxLayout
from PyQt5.QtGui import QIcon
from PyQt5 import QtCore


EXCELPATH = ""

# Start program# GUI
class XLWidget(QWidget):
    """
    Qt Widget for XL Reader.
    Contains an information label and a button to find your .exe file.
    """

    def __init__(self):
        super().__init__()

        self.initialize_gui()

    def initialize_gui(self):
        """
        Initialize the window.
        """
        self.setWindowTitle('XLReader')
        self.setWindowIcon(QIcon('icon.png'))

        self.InfoLabel = QLabel(self)
        self.InfoLabel.setText("Click the button below to find your Excel workbook.")
        self.InfoLabel.setStyleSheet(r"""QLabel { background-color : lightgray;
                                                 padding : 20px; }""")
        self.InfoLabel.setWordWrap(True)

        self.SelectXLButton = QPushButton(self)
        self.SelectXLButton.setText("Find Excel file")

        self.ProgressBar = QProgressBar(self)
        self.ProgressBar.setEnabled(False)

        layout = QVBoxLayout(self)
        layout.addWidget(self.InfoLabel)
        layout.addWidget(self.SelectXLButton)
        layout.addWidget(self.ProgressBar)

        self.SelectXLButton.clicked.connect(self.open_xl)

    def process_xl(self, file):
        """
        Takes a tuple holding
        a file with an absolute path and
        the type of the file.

        Processes the file, then saves the result to a new file
        with the name <filename>_orderedbyASINs.xlsx
        """
        
        self.ProgressBar.setValue(20)

        wbook = None

        try:
            wbookname = file[0]
            wbook = openpyxl.load_workbook(wbookname)
        except FileNotFoundError:
            self.InfoLabel.setText('File "' +
                                   wbookname +
                                   '" was not found. Check your spelling ' +
                                   'and capitalization and try again.')
            self.InfoLabel.setStyleSheet(r""" QLabel { color : red;
                                                       background-color : lightgray;
                                                       padding : 20px; }""")
            self.ProgressBar.setValue(0)
            self.ProgressBar.setEnabled(False)
            return

        if wbook is None:
            return

        asins = {}

        self.ProgressBar.setValue(40)

        # Open workbook
        try:
            wsheet1 = wbook.get_sheet_by_name('Sheet1')
            for row in range(2, wsheet1.max_row):
                asin_cell = "C" + str(row)
                asin = wsheet1[asin_cell].value
                quantity_cell = "E" + str(row)
                quantity = wsheet1[quantity_cell].value
                if asin not in asins.keys():  # New value
                    asins[asin] = list()
                    asins[asin].append(wsheet1["A" + str(row)].value)
                    asins[asin].append(wsheet1["B" + str(row)].value)
                    asins[asin].append(wsheet1["D" + str(row)].value)
                    asins[asin].append(quantity)
                    asins[asin].append(wsheet1["F" + str(row)].value)
                    asins[asin].append(wsheet1["G" + str(row)].value)
                else:
                    asins[asin][3] += quantity   # Add to old value
        except KeyError:
            self.InfoLabel.setText('Sheet "Sheet1" not found. Contact your administrator.')
            self.InfoLabel.setStyleSheet(r""" QLabel { color : red;
                                                       background-color : lightgray;
                                                       padding : 20px; }""")
            self.ProgressBar.setValue(0)
            self.ProgressBar.setEnabled(False)
            return

        # Save to new workbook
        newwbook = openpyxl.Workbook()
        wsheet = newwbook.active
        wsheet.title = 'CorrectedASINS'
        
        self.ProgressBar.setValue(60)

        wsheet["A1"] = "FC"
        wsheet["B1"] = "Category"
        wsheet["C1"] = "ASIN"
        wsheet["D1"] = "Description"
        wsheet["E1"] = "Units"
        wsheet["F1"] = "Cost"
        wsheet["G1"] = "Ext Cost"

        count = 2
        for asin in asins:
            wsheet["A" + str(count)] = asins[asin][0]
            wsheet["B" + str(count)] = asins[asin][1]
            wsheet["C" + str(count)] = asin
            wsheet["D" + str(count)] = asins[asin][2]
            wsheet["E" + str(count)] = asins[asin][3]
            wsheet["F" + str(count)] = asins[asin][4]
            wsheet["G" + str(count)] = asins[asin][5]
            count += 1

        self.ProgressBar.setValue(80)

        if '.xlsx' in wbookname:
            newwbookname = wbookname[:wbookname.rfind('.xlsx')] + '_orderedbyASINs' + ".xlsx"
        try:
            newwbook.save(newwbookname)
            self.InfoLabel.setText("File found. Excel workbook created in directory " +
                                   newwbookname[:newwbookname.rfind('/')] +
                                   " and named " +
                                   newwbookname[newwbookname.rfind('/') + 1:])
        except PermissionError:
            self.InfoLabel.setText('Could not save workbook. Please close ' +
                                   newwbookname +
                                   ', then try again.')
            self.ProgressBar.setValue(0)
            self.ProgressBar.setEnabled(False)

        self.ProgressBar.setValue(100)

    def open_xl(self):
        """
        Called on <self.btnSelectXLButton> click.

        Updates info label, accepts Excel workbook, and processes specified workbook.
        """
        self.ProgressBar.setEnabled(True)
        file = QFileDialog.getOpenFileName(self, "Open Excel")
        self.InfoLabel.setText('Loading workbook...')
        self.process_xl(file)

def main():
    """
    Main method for XLReader.
    """
    app = QApplication(sys.argv)
    window = XLWidget()
    window.setMinimumWidth(400)
    window.setMinimumHeight(300)
    window.show()

    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
